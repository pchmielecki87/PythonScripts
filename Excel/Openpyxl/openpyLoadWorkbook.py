from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# Load workbook
wb = load_workbook('./data/regions.xlsx')
new_sheet = wb.create_sheet('ImportedSheet')
active_sheet = wb.active

# Select cell
cell = active_sheet['A1']

# Print selected cell in active sheet
print(cell)

# Print selected cell value in active sheet
print(cell.value)

# Edit exising workbook and save as modified (folder must be in place)
active_sheet['A1'] = 0
wb.save('./data/modified/modified.xlsx')