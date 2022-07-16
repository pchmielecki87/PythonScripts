from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# Load workbook
wb = load_workbook('./data/regions.xlsx')
ws = wb.active

# Select cells, columns, rows
cell_range = ws['A1':'C1']
col_c = ws['C']
col_range = ws['A':'C']
row_range = ws[1:5]

# Print cell, column, row ranges
# print(cell_range)
# print(col_range)
# print(row_range)

# For
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
    for cell in row:
        print(cell)