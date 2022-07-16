from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# Create new workbook (Excel file)
wb = Workbook()

# Use active sheet and rename it
ws = wb.active
ws.title = 'My New Sheet'

# Create new Excel sheet (tab) and order them
ws1 = wb.create_sheet('NewSheet', 0)
ws2 = wb.create_sheet('AnotherSheet', 1)

# Print names of the sheets
print(wb.sheetnames)